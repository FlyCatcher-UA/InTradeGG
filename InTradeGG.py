import openpyxl



# читаем excel-файл
wb = openpyxl.load_workbook('/Users/PREDATOR/Desktop/dbInTradeGG.xlsx')

# печатаем список листов
# sheets = wb.sheetnames
# for sheet in sheets:
#     print(sheet)

# получаем активный лист
sheet = wb.active
rows = sheet.max_row
start=True

# перевірка структури файлу
if sheet.cell(row = 1, column = 2).value == "Название сделки" and sheet.cell(row = 1, column = 3).value == "Компания" and sheet.cell(row = 1, column = 7).value == "Этап сделки" and sheet.cell(row = 1, column = 132).value == "Назва товару 1" and sheet.cell(row = 1, column = 133).value == "Ціна за кг, євро" and sheet.cell(row = 1, column = 134).value == "Умови інкотермс" and sheet.cell(row = 1, column = 135).value == "Пройшла ціна?" and sheet.cell(row = 1, column = 137).value == "Назва товару 2" and sheet.cell(row = 1, column = 138).value == "Ціна за кг, євро" and sheet.cell(row = 1, column = 139).value == "Умови інкотермс" and sheet.cell(row = 1, column = 140).value == "Пройшла ціна?" and sheet.cell(row = 1, column = 142).value == "Назва товару 3" and sheet.cell(row = 1, column = 143).value == "Ціна за кг, євро" and sheet.cell(row = 1, column = 144).value == "Умови інкотермс" and sheet.cell(row = 1, column = 145).value == "Пройшла ціна?" :
    print("Структура файлу данних коректна")
    print("")
else:
    print("Структура ПОШКОДЖЕНА!")
    print("")


def first_stage():
    
    # робим уніфікований список компаній та присвоюєм їм ІД
    l_nouni_komp=[]
    l_uni_komp=[]
    
    
    for i in range (2, rows+1):
        cell = sheet.cell(row = i, column = 3)
        l_nouni_komp.append(cell.value)
        if cell.value not in l_uni_komp and cell.value != None:
            l_uni_komp.append(cell.value)
    
    # уніфікований список виводим
    l_uni_komp.sort()  
    for j in range(len(l_uni_komp)):
        print (f"{j+1} - {l_uni_komp[j]}")
    
    # рахуэм кількість сделок компанії    
    flag = True
    while flag == True:
        l_nomenklatur=[]
        l_price_nomenklatur=[]
        l_v_price_nomenklatur=[]
        l_data_close = []
        l_nice_price = []
        print("\n")    
        id_komp = input("Введіть ID компанії, або 0 для виходу: ")
        # перевірка чи ІД це числа
        if id_komp.isdigit() != True:
            print("\n"*2)
            print("----------------------------------------")        
            print("Такого ID немає")        
            continue
        else:
            id_komp = int(id_komp)
        # вихід з програми
        if id_komp == 0:
            flag = False
            break
        # перевірка чи ІД не виходить за рамки значень списка компаній
        if id_komp > len(l_uni_komp):
            print("\n"*2)
            print("----------------------------------------")        
            print("Такого ID немає")        
            continue
        else:
            print("\n"*2)
            print("----------------------------------------")
            count_sels=l_nouni_komp.count(l_uni_komp[id_komp-1])
            print(f"{l_uni_komp[id_komp-1]} - у моїй базі {count_sels} угод(а) для цієї компанії:")
            print("")
    
            # пошук деталей угод 
            for i in range(2, rows+1):
                cell = sheet.cell(row = i, column = 3)
                nomen = sheet.cell(row = i, column = 2)
                status = sheet.cell(row = i, column = 7)
                data_close = sheet.cell(row = i, column = 16)
                
                nomen1 = sheet.cell(row = i, column = 132)
                price1 = sheet.cell(row = i, column = 133)
                v_price1 = sheet.cell(row = i, column = 134)
                nice_price1 = sheet.cell(row = i, column = 135)
                
                nomen2 = sheet.cell(row = i, column = 137)
                price2 = sheet.cell(row = i, column = 138)
                v_price2 = sheet.cell(row = i, column = 139)
                nice_price2 = sheet.cell(row = i, column = 140)
                
                nomen3 = sheet.cell(row = i, column = 142)
                price3 = sheet.cell(row = i, column = 143)
                v_price3 = sheet.cell(row = i, column = 144)
                nice_price3 = sheet.cell(row = i, column = 145)
                
                # Опрацювання першої колонки номенклатур з виводом даних
                if cell.value == l_uni_komp[id_komp-1] and nomen1.value != None:
                    print(f"> {data_close.value} '{nomen.value}' - зі статусом {status.value} | {nomen1.value} заявлена ціна {price1.value} {v_price1.value} | {nice_price1.value}")
                    print("")
                elif cell.value == l_uni_komp[id_komp-1] and price1.value != None:
                    print(f"> {data_close.value} '{nomen.value}' - зі статусом {status.value} | заявлена ціна {price1.value} {v_price1.value} | {nice_price1.value}")
                    print("")
                elif cell.value == l_uni_komp[id_komp-1]:
                    print(f"> {data_close.value} '{nomen.value}' - зі статусом {status.value}")
                    print("")
                if cell.value == l_uni_komp[id_komp-1] and price1.value != None :
                    l_nomenklatur.append(nomen1.value.title())
                    l_price_nomenklatur.append(price1.value)
                    l_v_price_nomenklatur.append(v_price1.value)
                    l_data_close.append(data_close.value)
                    l_nice_price.append(nice_price1.value)
                    
                # Опрацювання другої колонки номенклатур з виводом даних
                if cell.value == l_uni_komp[id_komp-1] and nomen2.value != None:
                    print(f"> {data_close.value} '{nomen.value}' - зі статусом {status.value} | {nomen2.value} заявлена ціна {price2.value} {v_price2.value} | {nice_price2.value}")
                    print("")
                elif cell.value == l_uni_komp[id_komp-1] and price2.value != None:
                    print(f"> {data_close.value} '{nomen.value}' - зі статусом {status.value} | заявлена ціна {price2.value} {v_price2.value} | {nice_price2.value}")
                    print("")
                elif cell.value == l_uni_komp[id_komp-1]:
                    print(f"> {data_close.value} '{nomen.value}' - зі статусом {status.value}")
                    print("")
                if cell.value == l_uni_komp[id_komp-1] and price2.value != None:
                    l_nomenklatur.append(nomen2.value.title())
                    l_price_nomenklatur.append(price2.value)
                    l_v_price_nomenklatur.append(v_price2.value)
                    l_data_close.append(data_close.value)
                    l_nice_price.append(nice_price2.value)
                    
                # Опрацювання третьої колонки номенклатур з виводом даних
                if cell.value == l_uni_komp[id_komp-1] and nomen3.value != None:
                    print(f"> {data_close.value} '{nomen.value}' - зі статусом {status.value} | {nomen3.value} заявлена ціна {price3.value} {v_price3.value} | {nice_price3.value}")
                    print("")
                elif cell.value == l_uni_komp[id_komp-1] and price3.value != None:
                    print(f"> {data_close.value} '{nomen.value}' - зі статусом {status.value} | заявлена ціна {price3.value} {v_price3.value} | {nice_price3.value}")
                    print("")
                elif cell.value == l_uni_komp[id_komp-1]:
                    print(f"> {data_close.value} '{nomen.value}' - зі статусом {status.value}")
                    print("")
                if cell.value == l_uni_komp[id_komp-1] and price3.value != None:
                    l_nomenklatur.append(nomen3.value.title())
                    l_price_nomenklatur.append(price3.value)
                    l_v_price_nomenklatur.append(v_price3.value)
                    l_data_close.append(data_close.value)
                    l_nice_price.append(nice_price3.value)
                
        # пошук деталей номенклатур з угод
        if len(l_nomenklatur) > 0:
            print("\n")
            print("----------------------------------------")
            print("Товари на які було названо ціну:")
            print("")
        for i in range(len(l_nomenklatur)):        
            print(f"{i+1} - {l_nomenklatur[i]} {l_price_nomenklatur[i]} {l_v_price_nomenklatur[i]} {l_nice_price[i]}")
            print("")
        flag2 = True
        while flag2 and len(l_nomenklatur) > 0:
            
            l_price_for_analiz = []
            l_data_for_analiz = [] 
            l_nice_price_for_analiz = []
            
            l_price_for_analiz_FCA = []
            l_data_for_analiz_FCA = [] 
            l_nice_price_for_analiz_FCA = []       
                    
            id_nomenklatury = input("Щоб проаналізувати товар вкажіть його ID або введіть 0 для вибору іншої компанії: ")
            if id_nomenklatury == "0":
                flag2 = False
                break
            print("\n"*2)
            print("----------------------------------------")
            print("")
            
            # аналіз номенклатур
            if id_nomenklatury.isdigit() != True:
                print("Такого ID немає")
                continue
            else: 
                id_nomenklatury = int(id_nomenklatury)            
            if id_nomenklatury > len(l_nomenklatur):
                print("Такого ID немає")
                
            
            elif id_nomenklatury <= len(l_nomenklatur):
                for i in range(len(l_nomenklatur)):
                    if l_nomenklatur[i] == l_nomenklatur[id_nomenklatury-1] and (l_nice_price[i]=="Так" or l_nice_price[i]=="так" or l_nice_price[i]=="ТАК") and l_v_price_nomenklatur[i] == "DAP":
                        l_price_for_analiz.append(l_price_nomenklatur[i])
                        l_data_for_analiz.append(l_data_close[i])
                        l_nice_price_for_analiz.append(l_nice_price[i])
                    elif l_nomenklatur[i] == l_nomenklatur[id_nomenklatury-1] and (l_nice_price[i]=="Так" or l_nice_price[i]=="так" or l_nice_price[i]=="ТАК") and l_v_price_nomenklatur[i] == "FCA":
                        l_price_for_analiz_FCA.append(l_price_nomenklatur[i])
                        l_data_for_analiz_FCA.append(l_data_close[i])
                        l_nice_price_for_analiz_FCA.append(l_nice_price[i])
                                                  
                print(f"{l_nomenklatur[id_nomenklatury-1]}")
                print("")
                try:
                    print(f"Середня ціна - {round(sum(l_price_for_analiz)/len(l_price_for_analiz), 2)} DAP Остання ціна - {round(l_price_for_analiz[0], 2)} DAP")
                    print(f"Ціна з найбільшою ймовірністю проходу для {l_uni_komp[id_komp-1]} - {round(((sum(l_price_for_analiz)/len(l_price_for_analiz))+l_price_for_analiz[0])/2, 2)} DAP")
                    print("")
                except ZeroDivisionError:
                    print("Дані прохідних цін DAP відсутні.")
                    print("")
                try:
                    print(f"Середня ціна - {round(sum(l_price_for_analiz_FCA)/len(l_price_for_analiz_FCA), 2)} FCA Остання ціна - {round(l_price_for_analiz_FCA[0], 2)} FCA")
                    print(f"Ціна з найбільшою ймовірністю проходу для {l_uni_komp[id_komp-1]} - {round(((sum(l_price_for_analiz_FCA)/len(l_price_for_analiz_FCA))+l_price_for_analiz_FCA[0])/2, 2)} FCA")
                except ZeroDivisionError:
                    print("Дані прохідних цін FCA відсутні.")
                    print("")
                print("")

def second_stage():
    print("")  
    # робим уніфікований список товарів та присвоюєм їм ІД
    l_nouni_tovars=[]
    l_uni_tovars=[]
    
    
    
    for i in range (2, rows+1):
        cell = sheet.cell(row = i, column = 132)
        cell2 = sheet.cell(row = i, column = 137)
        cell3 = sheet.cell(row = i, column = 142)
        l_nouni_tovars.append(cell.value)
        if cell.value not in l_uni_tovars and cell.value != None and cell.value.title() not in l_uni_tovars:
            l_uni_tovars.append(cell.value.title())
        if cell2.value not in l_uni_tovars and cell2.value != None and cell2.value.title() not in l_uni_tovars:
            l_uni_tovars.append(cell2.value.title())
        if cell3.value not in l_uni_tovars and cell3.value != None and cell3.value.title() not in l_uni_tovars:
            l_uni_tovars.append(cell3.value.title())
    
    # уніфікований список виводим
    l_uni_tovars.sort()  
    for j in range(len(l_uni_tovars)):
        print (f"{j+1} - {l_uni_tovars[j]}")
    print("")
    
    
    flag3 = True
    while flag3:
        l_price_tovar_DAP=[]
        l_price_tovar_FCA=[]
        id_nomenklatury = input("Введіть ID номенклатури, або 0 для виходу: ")
        print("----------------------------------------")
        if id_nomenklatury == "0":
            flag3 = False
        elif id_nomenklatury.isdigit() and int(id_nomenklatury) <= len(l_uni_tovars):
            
            # Пошук деталей номенклатур
            for i in range(2, rows+1):
                kom = sheet.cell(row = i, column = 3)
                name_deal = sheet.cell(row = i, column = 2)
                status = sheet.cell(row = i, column = 7)
                data_close = sheet.cell(row = i, column = 16)
                        
                nomen1 = sheet.cell(row = i, column = 132)
                price1 = sheet.cell(row = i, column = 133)
                v_price1 = sheet.cell(row = i, column = 134)
                nice_price1 = sheet.cell(row = i, column = 135)
                
                nomen2 = sheet.cell(row = i, column = 137)
                price2 = sheet.cell(row = i, column = 138)
                v_price2 = sheet.cell(row = i, column = 139)
                nice_price2 = sheet.cell(row = i, column = 140)
                
                nomen3 = sheet.cell(row = i, column = 142)
                price3 = sheet.cell(row = i, column = 143)
                v_price3 = sheet.cell(row = i, column = 144)
                nice_price3 = sheet.cell(row = i, column = 145)                
            
                if nomen1.value != None and nomen1.value.title() == l_uni_tovars[int(id_nomenklatury)-1]:
                    print(data_close.value, nomen1.value.title(), kom.value, price1.value, v_price1.value, nice_price1.value)
                    print("")                    
                    if v_price1.value == "DAP" and (nice_price1.value == "Так" or nice_price1.value == "так"  or nice_price1.value == "ТАК"):
                        l_price_tovar_DAP.append(price1.value)                        
                    if v_price1.value == "FCA" and (nice_price1.value == "Так" or nice_price1.value == "так"  or nice_price1.value == "ТАК"):
                        l_price_tovar_FCA.append(price1.value)
                        
                        
                if nomen2.value != None and nomen2.value.title() == l_uni_tovars[int(id_nomenklatury)-1]:
                    print(data_close.value, nomen2.value.title(), kom.value, price2.value, v_price2.value, nice_price2.value)
                    print("")                    
                    if v_price2.value == "DAP" and (nice_price2.value == "Так" or nice_price2.value == "так"  or nice_price2.value == "ТАК"):
                        l_price_tovar_DAP.append(price2.value)                        
                    if v_price2.value == "FCA" and (nice_price2.value == "Так" or nice_price2.value == "так"  or nice_price2.value == "ТАК"):
                        l_price_tovar_FCA.append(price2.value)
                        
                        
                if nomen3.value != None and nomen3.value.title() == l_uni_tovars[int(id_nomenklatury)-1]:
                    print(data_close.value, nomen3.value.title(), kom.value, price3.value, v_price3.value, nice_price3.value)
                    print("")                    
                    if v_price3.value == "DAP" and (nice_price3.value == "Так" or nice_price3.value == "так"  or nice_price3.value == "ТАК"):
                        l_price_tovar_DAP.append(price3.value)                        
                    if v_price3.value == "FCA" and (nice_price3.value == "Так" or nice_price3.value == "так"  or nice_price3.value == "ТАК"):
                        l_price_tovar_FCA.append(price3.value)
                        
            
            if len(l_price_tovar_DAP) > 0:
                print("----------------------------------------")
                print(f"Середня ціна DAP за весь період - {round(sum(l_price_tovar_DAP)/len(l_price_tovar_DAP), 2)}")
                print(f"Максимальна ціна DAP яка пройшла - {max(l_price_tovar_DAP)}")
                print(f"Мінімальна ціна DAP яка пройшла - {min(l_price_tovar_DAP)}")
                print(f"Остання ціна DAP яка пройшла - {l_price_tovar_DAP[0]}")
                print("")
            if len(l_price_tovar_FCA) > 0:
                print("----------------------------------------")
                print(f"Середня ціна FCA за весь період - {round(sum(l_price_tovar_FCA)/len(l_price_tovar_FCA), 2)}")
                print(f"Максимальна ціна FCA яка пройшла - {max(l_price_tovar_FCA)}")
                print(f"Мінімальна ціна FCA яка пройшла - {min(l_price_tovar_FCA)}")
                print(f"Остання ціна FCA яка пройшла - {l_price_tovar_FCA[0]}")
                print("")
        else:
            print("Такого ID немає.")
                
def enother_stage(): 
    print("Такого методу немає.")
    
while start:
    #Вибір режиму роботи з компаніями чи номенклатурами
    print("----------------------------------------")
    ans=input("Оберіть метод роботи: 1 - По компаніях, 2 - По товарах. 0 - Вихід: ")
    print("----------------------------------------")
    if ans == "1":
        first_stage()
    elif ans == "2":
        second_stage()
    elif ans == "0":
        start=False
    else:
        enother_stage()